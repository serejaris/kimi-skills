#!/usr/bin/env python3
"""
Excel Screenshot Tool - Export all sheets in an Excel file to PNG images
Uses HTML rendering with Playwright for accurate Excel styling.
Automatically segments large sheets into multiple screenshots for clarity.

IMPORTANT - Timeout:
    This script can take a long time for large Excel files. Each sheet segment
    takes ~2-5 seconds to render. A complex workbook with 500+ rows across
    multiple sheets can easily take 2-5 minutes total.

    When calling this script, set a generous subprocess timeout:
      - Small files (1-2 sheets, <100 rows):  ~30 seconds
      - Medium files (3-5 sheets, <300 rows):  ~120 seconds
      - Large files (5+ sheets, 500+ rows):    ~300 seconds (5 minutes)
      - Very large files:                      ~600 seconds (10 minutes)

    Example with timeout:
      subprocess.run(["python", "excel_screenshot.py", "input.xlsx", "output/"],
                     timeout=300)

    Do NOT use the default 30s or 60s timeout for this script. If the file is
    large, the script is still working correctly - it just needs more time.

Usage:
    python excel_screenshot.py <input.xlsx> <output_directory>
    python excel_screenshot.py <input.xlsx> <output_directory> --max-rows 30

Output:
    Creates PNG images for each sheet (or sheet segment) in the output directory.
    Large sheets are automatically split into segments of --max-rows rows each.

Dependencies:
    pip install openpyxl Pillow playwright && playwright install chromium
"""

import os
import sys
import re
import argparse
import base64
import time
import zipfile
import xml.etree.ElementTree as ET
from pathlib import Path
from dataclasses import dataclass, field
from typing import Optional, Callable, Dict, Tuple

from openpyxl import load_workbook
from openpyxl.cell.cell import Cell, MergedCell
from openpyxl.utils import get_column_letter

from PIL import Image as PILImage, ImageChops


# ============================================================
# Config
# ============================================================

MAX_ROWS_PER_SEGMENT = 30
MAX_COLS_FOR_SCREENSHOT = 100
SHEET_RENDER_TIMEOUT = 90  # seconds per segment


# ============================================================
# Auto-segmentation
# ============================================================

def auto_segment(sheet, max_rows: int = MAX_ROWS_PER_SEGMENT) -> list:
    """Split a sheet into segments of at most max_rows rows.
    Tries to break at empty rows near the boundary.
    Returns list of (start_row, end_row).
    """
    total = sheet.max_row or 1
    max_col = sheet.max_column or 1
    segments = []
    start = 1

    while start <= total:
        end = min(start + max_rows - 1, total)
        if end >= total:
            segments.append((start, total))
            break

        # Look for an empty row near the boundary (within 5 rows)
        best_break = end
        for probe in range(end, max(start, end - 5) - 1, -1):
            if all(sheet.cell(row=probe, column=c).value is None
                   for c in range(1, max_col + 1)):
                best_break = probe
                break

        segments.append((start, best_break))
        start = best_break + 1

    return segments


# ============================================================
# Dataclasses
# ============================================================

@dataclass
class SheetScreenshot:
    sheet_name: str
    output_file: str = ""
    success: bool = False
    error: Optional[str] = None


@dataclass
class ScreenshotResult:
    status: str = "success"
    input_file: str = ""
    output_directory: str = ""
    total_sheets: int = 0
    success_count: int = 0
    failed_count: int = 0
    sheets: list = field(default_factory=list)
    error: Optional[str] = None


# ============================================================
# Utility functions
# ============================================================

def sanitize_filename(name: str) -> str:
    """Sanitize filename to remove invalid characters"""
    sanitized = re.sub(r'[<>:"/\\|?*\x00-\x1f]', '_', name)
    sanitized = sanitized.strip()
    if not sanitized:
        sanitized = "sheet"
    if len(sanitized) > 100:
        sanitized = sanitized[:100]
    return sanitized


def get_output_directory(input_name: str, output_base: str) -> str:
    """Get output directory under the given base path."""
    return os.path.join(output_base, input_name)


def auto_crop_whitespace(png_file: str):
    """Auto-crop whitespace from a PNG file using Pillow."""
    try:
        img = PILImage.open(png_file)
        bg = PILImage.new(img.mode, img.size, (255, 255, 255))
        diff = ImageChops.difference(img, bg)
        bbox = diff.getbbox()
        if bbox:
            pad = 4
            x0 = max(0, bbox[0] - pad)
            y0 = max(0, bbox[1] - pad)
            x1 = min(img.width, bbox[2] + pad)
            y1 = min(img.height, bbox[3] + pad)
            img.crop((x0, y0, x1, y1)).save(png_file)
    except Exception:
        pass


# ============================================================
# Excel color extraction
# ============================================================

def apply_tint(hex_color: str, tint: float) -> str:
    """Apply tint to a color (Excel's tint algorithm)"""
    if not hex_color or len(hex_color) < 6:
        return hex_color
    if len(hex_color) == 8:
        hex_color = hex_color[2:]
    try:
        r = int(hex_color[0:2], 16)
        g = int(hex_color[2:4], 16)
        b = int(hex_color[4:6], 16)
    except ValueError:
        return hex_color

    if tint < 0:
        r = int(r * (1 + tint))
        g = int(g * (1 + tint))
        b = int(b * (1 + tint))
    elif tint > 0:
        r = int(r + (255 - r) * tint)
        g = int(g + (255 - g) * tint)
        b = int(b + (255 - b) * tint)

    r, g, b = max(0, min(255, r)), max(0, min(255, g)), max(0, min(255, b))
    return f"{r:02X}{g:02X}{b:02X}"


def get_theme_color(wb, theme_index: int, tint: float = 0) -> Optional[str]:
    """Get theme color from workbook"""
    try:
        if hasattr(wb, 'loaded_theme') and wb.loaded_theme:
            color_map = {
                0: 'FFFFFF', 1: '000000', 2: 'EEECE1', 3: '1F497D',
                4: '4F81BD', 5: 'C0504D', 6: '9BBB59', 7: '8064A2',
                8: '4BACC6', 9: 'F79646',
            }
            if theme_index in color_map:
                return apply_tint(color_map[theme_index], tint)
    except Exception:
        pass
    return None


def get_cell_color(color_obj, wb) -> Optional[str]:
    """Extract color from openpyxl color object"""
    if color_obj is None:
        return None

    if color_obj.type == 'rgb' and color_obj.rgb:
        rgb = color_obj.rgb
        if isinstance(rgb, str) and len(rgb) >= 6:
            if rgb == "00000000":
                return None
            if len(rgb) == 8:
                rgb = rgb[2:]
            return rgb

    if color_obj.type == 'theme' and color_obj.theme is not None:
        tint = color_obj.tint if color_obj.tint else 0
        return get_theme_color(wb, color_obj.theme, tint)

    if color_obj.type == 'indexed' and color_obj.indexed is not None:
        indexed_colors = {
            0: '000000', 1: 'FFFFFF', 2: 'FF0000', 3: '00FF00',
            4: '0000FF', 5: 'FFFF00', 6: 'FF00FF', 7: '00FFFF',
            8: '000000', 9: 'FFFFFF', 10: 'FF0000', 11: '00FF00',
            12: '0000FF', 13: 'FFFF00', 14: 'FF00FF', 15: '00FFFF',
            16: '800000', 17: '008000', 18: '000080', 19: '808000',
            20: '800080', 21: '008080', 22: 'C0C0C0', 23: '808080',
            24: '9999FF', 25: '993366', 26: 'FFFFCC', 27: 'CCFFFF',
            28: '660066', 29: 'FF8080', 30: '0066CC', 31: 'CCCCFF',
            32: '000080', 33: 'FF00FF', 34: 'FFFF00', 35: '00FFFF',
            36: '800080', 37: '800000', 38: '008080', 39: '0000FF',
            40: '00CCFF', 41: 'CCFFFF', 42: 'CCFFCC', 43: 'FFFF99',
            44: '99CCFF', 45: 'FF99CC', 46: 'CC99FF', 47: 'FFCC99',
            48: '3366FF', 49: '33CCCC', 50: '99CC00', 51: 'FFCC00',
            52: 'FF9900', 53: 'FF6600', 54: '666699', 55: '969696',
            56: '003366', 57: '339966', 58: '003300', 59: '333300',
            60: '993300', 61: '993366', 62: '333399', 63: '333333',
            64: None, 65: None,
        }
        if color_obj.indexed in indexed_colors:
            return indexed_colors[color_obj.indexed]

    return None


# ============================================================
# Cell style extraction
# ============================================================

def get_cell_style(cell: Cell, wb) -> str:
    """Get CSS styles for a cell"""
    if isinstance(cell, MergedCell):
        return ""
    styles = []

    if cell.fill and cell.fill.patternType and cell.fill.patternType != 'none':
        fg_color = get_cell_color(cell.fill.fgColor, wb)
        if fg_color:
            styles.append(f"background-color: #{fg_color}")

    if cell.font and cell.font.color:
        font_color = get_cell_color(cell.font.color, wb)
        if font_color and font_color != "000000":
            styles.append(f"color: #{font_color}")

    if cell.font and cell.font.bold:
        styles.append("font-weight: bold")
    if cell.font and cell.font.italic:
        styles.append("font-style: italic")
    if cell.font and cell.font.size:
        styles.append(f"font-size: {int(cell.font.size * 1.33)}px")

    if cell.alignment:
        if cell.alignment.horizontal in ('left', 'center', 'right', 'justify'):
            styles.append(f"text-align: {cell.alignment.horizontal}")
        v = cell.alignment.vertical
        if v == "center":
            styles.append("vertical-align: middle")
        elif v in ("top", "bottom"):
            styles.append(f"vertical-align: {v}")
        if cell.alignment.wrap_text:
            styles.append("white-space: pre-wrap")

    if cell.border:
        style_map = {
            'thin': '1px solid', 'medium': '2px solid', 'thick': '3px solid',
            'dashed': '1px dashed', 'dotted': '1px dotted', 'double': '3px double',
            'hair': '1px solid', 'mediumDashed': '2px dashed',
            'dashDot': '1px dashed', 'mediumDashDot': '2px dashed',
            'dashDotDot': '1px dashed', 'mediumDashDotDot': '2px dashed',
            'slantDashDot': '2px dashed',
        }
        for side in ('left', 'right', 'top', 'bottom'):
            b = getattr(cell.border, side, None)
            if b and b.style and b.style != 'none':
                css = style_map.get(b.style, '1px solid')
                color = '#999'
                if b.color:
                    c = get_cell_color(b.color, wb)
                    if c:
                        color = f'#{c}'
                styles.append(f"border-{side}: {css} {color}")

    return "; ".join(styles)


def format_cell_value(value) -> str:
    """Format cell value for HTML display"""
    if value is None:
        return ""
    if isinstance(value, bool):
        return "TRUE" if value else "FALSE"
    if isinstance(value, float):
        return f"{int(value):,}" if value == int(value) else f"{value:,.2f}"
    if isinstance(value, int):
        return f"{value:,}"
    text = str(value)
    text = text.replace("&", "&amp;").replace("<", "&lt;").replace(">", "&gt;").replace("\n", "<br>")
    return text


# ============================================================
# Column / Row dimensions
# ============================================================

def get_column_widths(sheet) -> dict:
    widths = {}
    for col_letter, col_dim in sheet.column_dimensions.items():
        if col_dim.width:
            widths[col_letter] = int(col_dim.width * 7)
    return widths


def get_row_heights(sheet) -> dict:
    heights = {}
    for row_num, row_dim in sheet.row_dimensions.items():
        if row_dim.height:
            heights[row_num] = int(row_dim.height * 1.33)
    return heights


# ============================================================
# Embedded image extraction (richData / WPS DISPIMG / floating)
# ============================================================

def extract_richdata_images(xlsx_path: str) -> Dict[int, dict]:
    """Extract images embedded via Excel 365 richData (IMAGE function)."""
    richdata_images = {}
    try:
        with zipfile.ZipFile(xlsx_path, 'r') as z:
            rels_path = 'xl/richData/_rels/richValueRel.xml.rels'
            if rels_path not in z.namelist():
                return richdata_images

            rels_content = z.read(rels_path).decode('utf-8')
            rels_root = ET.fromstring(rels_content)

            for rel in rels_root.iter():
                if 'Relationship' in rel.tag:
                    rel_id = rel.get('Id', '')
                    rel_type = rel.get('Type', '')
                    target = rel.get('Target', '')

                    if 'image' in rel_type.lower() and target:
                        if target.startswith('..'):
                            img_path = 'xl/' + target[3:]
                        else:
                            img_path = 'xl/richData/' + target

                        try:
                            idx = int(rel_id.replace('rId', '')) - 1
                        except Exception:
                            idx = 0

                        if img_path in z.namelist():
                            raw_data = z.read(img_path)
                            img_data = base64.b64encode(raw_data).decode('utf-8')

                            img_format = 'png'
                            if raw_data[:2] == b'\xff\xd8':
                                img_format = 'jpeg'
                            elif raw_data[:6] in (b'GIF87a', b'GIF89a'):
                                img_format = 'gif'

                            richdata_images[idx] = {
                                'data': img_data,
                                'format': img_format,
                                'path': img_path,
                            }
    except Exception:
        pass
    return richdata_images


def extract_wps_cellimages(xlsx_path: str) -> Dict[str, dict]:
    """Extract images embedded via WPS DISPIMG function."""
    wps_images = {}
    try:
        with zipfile.ZipFile(xlsx_path, 'r') as z:
            if 'xl/cellimages.xml' not in z.namelist():
                return wps_images

            rels_path = 'xl/_rels/cellimages.xml.rels'
            if rels_path not in z.namelist():
                return wps_images

            rels_content = z.read(rels_path).decode('utf-8')
            rels_root = ET.fromstring(rels_content)

            rid_to_path = {}
            for rel in rels_root.iter():
                if 'Relationship' in rel.tag:
                    rel_id = rel.get('Id', '')
                    target = rel.get('Target', '')
                    if target:
                        if target.startswith('media/'):
                            img_path = 'xl/' + target
                        else:
                            img_path = 'xl/media/' + target
                        rid_to_path[rel_id] = img_path

            cellimages_content = z.read('xl/cellimages.xml').decode('utf-8')
            cellimages_root = ET.fromstring(cellimages_content)

            for elem in cellimages_root.iter():
                if 'cNvPr' in elem.tag:
                    img_id = elem.get('name', '')
                    if img_id and img_id.startswith('ID_'):
                        for pic in cellimages_root.iter():
                            if 'pic' in pic.tag.lower():
                                for child in pic.iter():
                                    if 'cNvPr' in child.tag and child.get('name') == img_id:
                                        for blip in pic.iter():
                                            if 'blip' in blip.tag.lower():
                                                for attr_name, attr_val in blip.attrib.items():
                                                    if 'embed' in attr_name.lower():
                                                        rid = attr_val
                                                        if rid in rid_to_path:
                                                            img_path = rid_to_path[rid]
                                                            if img_path in z.namelist():
                                                                raw_data = z.read(img_path)
                                                                img_data = base64.b64encode(raw_data).decode('utf-8')
                                                                img_format = 'jpeg'
                                                                if raw_data[:8] == b'\x89PNG\r\n\x1a\n':
                                                                    img_format = 'png'
                                                                elif raw_data[:2] == b'\xff\xd8':
                                                                    img_format = 'jpeg'
                                                                elif raw_data[:6] in (b'GIF87a', b'GIF89a'):
                                                                    img_format = 'gif'
                                                                wps_images[img_id] = {
                                                                    'data': img_data,
                                                                    'format': img_format,
                                                                    'path': img_path,
                                                                }
                                                        break
                                        break
    except Exception:
        pass
    return wps_images


def extract_dispimg_cells(xlsx_path: str, sheet_index: int = 0) -> Dict[Tuple[int, int], str]:
    """Extract mapping of cells containing DISPIMG function to their image IDs."""
    dispimg_cells = {}
    try:
        with zipfile.ZipFile(xlsx_path, 'r') as z:
            sheet_file = f'xl/worksheets/sheet{sheet_index + 1}.xml'
            if sheet_file not in z.namelist():
                return dispimg_cells

            sheet_content = z.read(sheet_file).decode('utf-8')
            sheet_content = sheet_content.replace('&quot;', '"').replace('&amp;', '&').replace('&lt;', '<').replace('&gt;', '>')

            cell_pattern = r'<c r="([A-Z]+)(\d+)"[^>]*(?<!/)>(.*?)</c>'
            for match in re.finditer(cell_pattern, sheet_content):
                col_str = match.group(1)
                row_str = match.group(2)
                cell_content = match.group(3)

                if 'DISPIMG' not in cell_content.upper():
                    continue

                id_match = re.search(r'DISPIMG\s*\(\s*"(ID_[A-F0-9]+)"', cell_content, re.IGNORECASE)
                if not id_match:
                    continue

                img_id = id_match.group(1)
                col = 0
                for c in col_str.upper():
                    col = col * 26 + (ord(c) - ord('A') + 1)
                row = int(row_str)
                dispimg_cells[(row, col)] = img_id
    except Exception:
        pass
    return dispimg_cells


def extract_cell_richdata_mapping(xlsx_path: str, sheet_name: str) -> Dict[Tuple[int, int], int]:
    """Extract mapping of cells to richData value metadata indices."""
    cell_vm_map = {}
    try:
        with zipfile.ZipFile(xlsx_path, 'r') as z:
            sheet_file = None
            for i, name in enumerate(['sheet1.xml', 'sheet2.xml', 'sheet3.xml', 'sheet4.xml', 'sheet5.xml']):
                path = f'xl/worksheets/{name}'
                if path in z.namelist():
                    if i == 0:
                        sheet_file = path
                        break

            if not sheet_file:
                return cell_vm_map

            sheet_content = z.read(sheet_file).decode('utf-8')
            root = ET.fromstring(sheet_content)

            for elem in root.iter():
                if elem.tag.endswith('}c') or elem.tag == 'c':
                    vm = elem.get('vm')
                    r = elem.get('r')
                    if vm and r:
                        col_str = ''.join(c for c in r if c.isalpha())
                        row_str = ''.join(c for c in r if c.isdigit())
                        if col_str and row_str:
                            col = 0
                            for c in col_str.upper():
                                col = col * 26 + (ord(c) - ord('A') + 1)
                            row = int(row_str)
                            cell_vm_map[(row, col)] = int(vm) - 1
    except Exception:
        pass
    return cell_vm_map


def extract_images_from_sheet(sheet, col_widths: dict, row_heights: dict) -> list:
    """Extract floating images from a sheet with pixel positions."""
    floating_images = []
    if not hasattr(sheet, '_images') or not sheet._images:
        return floating_images

    for img in sheet._images:
        try:
            anchor = img.anchor
            if not hasattr(anchor, '_from'):
                continue
            from_cell = anchor._from
            col, row = from_cell.col, from_cell.row

            pixel_x = 10
            for c in range(1, col + 1):
                pixel_x += col_widths.get(get_column_letter(c), 64)
            if hasattr(from_cell, 'colOff') and from_cell.colOff:
                pixel_x += int(from_cell.colOff / 914400 * 96)

            pixel_y = 10
            for r in range(1, row + 1):
                pixel_y += row_heights.get(r, 20)
            if hasattr(from_cell, 'rowOff') and from_cell.rowOff:
                pixel_y += int(from_cell.rowOff / 914400 * 96)

            width = img.width if hasattr(img, 'width') else 100
            height = img.height if hasattr(img, 'height') else 100
            if width > 100000:
                width = int(width / 914400 * 96)
            if height > 100000:
                height = int(height / 914400 * 96)

            raw_data = None
            if hasattr(img, '_data') and callable(img._data):
                try:
                    raw_data = img._data()
                except Exception:
                    pass
            if raw_data is None and hasattr(img, 'ref') and img.ref:
                try:
                    if hasattr(img.ref, 'read'):
                        img.ref.seek(0)
                        raw_data = img.ref.read()
                    elif hasattr(img.ref, 'blob'):
                        raw_data = img.ref.blob
                except Exception:
                    pass

            if raw_data:
                img_format = 'png'
                if raw_data[:2] == b'\xff\xd8':
                    img_format = 'jpeg'
                elif raw_data[:6] in (b'GIF87a', b'GIF89a'):
                    img_format = 'gif'
                elif raw_data[:4] == b'RIFF' and raw_data[8:12] == b'WEBP':
                    img_format = 'webp'

                floating_images.append({
                    'x': pixel_x, 'y': pixel_y,
                    'width': max(width, 20), 'height': max(height, 20),
                    'data': base64.b64encode(raw_data).decode('utf-8'),
                    'format': img_format,
                })
        except Exception:
            continue
    return floating_images


# ============================================================
# HTML generation (supports row range for segmentation)
# ============================================================

def sheet_to_html(sheet, sheet_name: str, wb, xlsx_path: str = None,
                  richdata_images: Dict[int, dict] = None,
                  cell_vm_map: Dict[Tuple[int, int], int] = None,
                  wps_images: Dict[str, dict] = None,
                  dispimg_cells: Dict[Tuple[int, int], str] = None,
                  start_row: int = 1, end_row: int = None,
                  show_title: bool = True) -> str:
    """Convert a sheet (or a row range) to an HTML table.

    Args:
        start_row: First row to render (1-indexed, inclusive).
        end_row:   Last row to render (1-indexed, inclusive). None = sheet max.
        show_title: Whether to include <h2> sheet name header.
    """
    col_widths = get_column_widths(sheet)
    row_heights = get_row_heights(sheet)

    floating_images = extract_images_from_sheet(sheet, col_widths, row_heights)

    if richdata_images is None:
        richdata_images = {}
    if cell_vm_map is None:
        cell_vm_map = {}
    if wps_images is None:
        wps_images = {}
    if dispimg_cells is None:
        dispimg_cells = {}

    # Merged cells
    merged_cells = {}
    for merged_range in sheet.merged_cells.ranges:
        min_col, min_row, max_col, max_row = merged_range.bounds
        merged_cells[(min_row, min_col)] = (max_row - min_row + 1, max_col - min_col + 1)
        for row in range(min_row, max_row + 1):
            for col in range(min_col, max_col + 1):
                if (row, col) != (min_row, min_col):
                    merged_cells[(row, col)] = None

    raw_max_col = sheet.max_column or 1

    # Build visible column list: skip hidden columns, cap at MAX_COLS_FOR_SCREENSHOT
    visible_cols = []
    for c in range(1, raw_max_col + 1):
        letter = get_column_letter(c)
        dim = sheet.column_dimensions.get(letter)
        if dim and dim.hidden:
            continue
        visible_cols.append(c)
        if len(visible_cols) >= MAX_COLS_FOR_SCREENSHOT:
            break

    if end_row is None:
        end_row = sheet.max_row or 1

    # Column group
    colgroup_html = "<colgroup>\n"
    for col_idx in visible_cols:
        w = col_widths.get(get_column_letter(col_idx), 64)
        colgroup_html += f'    <col style="width: {w}px;">\n'
    colgroup_html += "</colgroup>\n"

    # Floating images HTML
    floating_images_html = ""
    for img_info in floating_images:
        floating_images_html += (
            f'    <img class="floating-image" '
            f'src="data:image/{img_info["format"]};base64,{img_info["data"]}" '
            f'style="left:{img_info["x"]}px;top:{img_info["y"]}px;'
            f'width:{img_info["width"]}px;height:{img_info["height"]}px;">\n'
        )

    title_html = f"    <h2>{sheet_name}</h2>\n" if show_title else ""

    html = f"""<!DOCTYPE html>
<html>
<head>
    <meta charset="UTF-8">
    <style>
        body {{
            font-family: 'Microsoft YaHei', 'SimHei', 'WenQuanYi Micro Hei', 'Hiragino Sans GB', 'Noto Sans CJK SC', 'PingFang SC', sans-serif;
            margin: 10px;
            padding: 0;
            background-color: white;
            position: relative;
        }}
        h2 {{
            color: #333;
            margin: 0 0 10px 0;
            font-size: 16px;
        }}
        table {{
            border-collapse: collapse;
            width: auto;
            font-size: 11px;
            table-layout: fixed;
        }}
        td {{
            border: 1px solid #d0d0d0;
            padding: 3px 6px;
            white-space: nowrap;
            overflow: hidden;
            text-overflow: ellipsis;
            vertical-align: middle;
        }}
        td.number {{
            text-align: right;
        }}
        .floating-image {{
            position: absolute;
            z-index: 10;
        }}
    </style>
</head>
<body>
{title_html}{floating_images_html}    <table>
{colgroup_html}
"""

    for row_idx in range(start_row, end_row + 1):
        row_height = row_heights.get(row_idx, 20)
        html += f'        <tr style="height: {row_height}px;">\n'

        for col_idx in visible_cols:
            cell_key = (row_idx, col_idx)

            if cell_key in merged_cells:
                if merged_cells[cell_key] is None:
                    continue

            cell = sheet.cell(row=row_idx, column=col_idx)

            span_attrs = ""
            if cell_key in merged_cells and merged_cells[cell_key] is not None:
                rowspan, colspan = merged_cells[cell_key]
                if rowspan > 1:
                    span_attrs += f' rowspan="{rowspan}"'
                if colspan > 1:
                    span_attrs += f' colspan="{colspan}"'

            value = cell.value
            display_value = None

            # Priority 1: Excel 365 richData images
            if cell_key in cell_vm_map:
                vm_idx = cell_vm_map[cell_key]
                if vm_idx in richdata_images:
                    img_info = richdata_images[vm_idx]
                    col_letter = get_column_letter(col_idx)
                    cell_width = col_widths.get(col_letter, 64) - 12
                    cell_height = row_heights.get(row_idx, 20) - 6
                    display_value = (
                        f'<img src="data:image/{img_info["format"]};base64,{img_info["data"]}" '
                        f'style="max-width:{cell_width}px;max-height:{cell_height}px;'
                        f'width:auto;height:auto;display:block;">'
                    )

            # Priority 2: WPS DISPIMG images
            if display_value is None and cell_key in dispimg_cells:
                img_id = dispimg_cells[cell_key]
                if img_id in wps_images:
                    img_info = wps_images[img_id]
                    col_letter = get_column_letter(col_idx)
                    cell_width = col_widths.get(col_letter, 64) - 12
                    cell_height = row_heights.get(row_idx, 20) - 6
                    display_value = (
                        f'<img src="data:image/{img_info["format"]};base64,{img_info["data"]}" '
                        f'style="max-width:{cell_width}px;max-height:{cell_height}px;'
                        f'width:auto;height:auto;display:block;">'
                    )

            # Priority 3: Filter out error values and DISPIMG formula text
            if display_value is None:
                if isinstance(value, str) and (
                    value.startswith('#') or
                    'DISPIMG' in value.upper() or
                    value.startswith('=DISPIMG')
                ):
                    display_value = ""
                else:
                    display_value = format_cell_value(value)

            cell_style = get_cell_style(cell, wb)
            style_attr = f' style="{cell_style}"' if cell_style else ""

            class_attr = ""
            if isinstance(value, (int, float)) and not isinstance(value, bool):
                class_attr = ' class="number"'

            html += f'            <td{span_attrs}{style_attr}{class_attr}>{display_value}</td>\n'

        html += '        </tr>\n'

    html += """    </table>
</body>
</html>
"""
    return html


# ============================================================
# HTML to Image Conversion
# ============================================================

def get_html_to_image_converter() -> Callable:
    """Get the best available HTML to image converter"""
    try:
        from playwright.sync_api import sync_playwright
        print("Using Playwright for HTML to image conversion")
        return convert_html_to_image_playwright
    except ImportError:
        pass

    try:
        import imgkit
        imgkit.from_string("<html></html>", False, options={'quiet': ''})
        print("Using imgkit (wkhtmltoimage) for HTML to image conversion")
        return convert_html_to_image_imgkit
    except Exception:
        pass

    raise RuntimeError(
        "No HTML to image converter available. Please install one of:\n"
        "  - Playwright: pip install playwright && playwright install chromium\n"
        "  - wkhtmltopdf: brew install wkhtmltopdf (macOS) or apt install wkhtmltopdf (Linux)"
    )


def convert_html_to_image_playwright(html_file: str, output_file: str, browser=None) -> bool:
    """Convert HTML to PNG using Playwright with 2x resolution and auto-crop."""
    from playwright.sync_api import sync_playwright

    timeout_ms = SHEET_RENDER_TIMEOUT * 1000

    def _render(br):
        page = br.new_page(viewport={"width": 1920, "height": 1080}, device_scale_factor=2)
        page.set_default_timeout(timeout_ms)
        try:
            page.goto(f"file://{os.path.abspath(html_file)}", timeout=timeout_ms)
            page.wait_for_load_state("networkidle", timeout=30000)

            dimensions = page.evaluate("""() => {
                const body = document.body;
                const html = document.documentElement;
                const width = Math.max(
                    body.scrollWidth, body.offsetWidth,
                    html.clientWidth, html.scrollWidth, html.offsetWidth
                );
                const height = Math.max(
                    body.scrollHeight, body.offsetHeight,
                    html.clientHeight, html.scrollHeight, html.offsetHeight
                );
                return { width, height };
            }""")
            page.set_viewport_size({
                "width": dimensions["width"] + 20,
                "height": dimensions["height"] + 20,
            })
            page.screenshot(path=output_file, full_page=True, type="png", timeout=timeout_ms)
        finally:
            page.close()

        # Auto-crop whitespace
        auto_crop_whitespace(output_file)

        return os.path.exists(output_file)

    if browser is not None:
        return _render(browser)

    with sync_playwright() as p:
        br = p.chromium.launch()
        try:
            return _render(br)
        finally:
            br.close()


def convert_html_to_image_imgkit(html_file: str, output_file: str, browser=None) -> bool:
    """Convert HTML to PNG using imgkit (wkhtmltoimage). Fallback when Playwright is unavailable."""
    import imgkit
    options = {
        'format': 'png',
        'encoding': 'UTF-8',
        'quality': '100',
        'width': '1400',
        'enable-local-file-access': None,
    }
    imgkit.from_file(html_file, output_file, options=options)
    return os.path.exists(output_file)


# ============================================================
# Main Screenshot Logic
# ============================================================

def take_screenshots(input_file: str, output_dir: str,
                     max_rows: int = MAX_ROWS_PER_SEGMENT) -> ScreenshotResult:
    """Take screenshots of all sheets in an Excel file, auto-segmenting large sheets."""
    result = ScreenshotResult(
        input_file=os.path.basename(input_file),
        output_directory=output_dir,
    )

    try:
        os.makedirs(output_dir, exist_ok=True)
    except Exception as e:
        result.status = "error"
        result.error = f"Failed to create output directory: {e}"
        return result

    try:
        convert_to_image = get_html_to_image_converter()
    except RuntimeError as e:
        result.status = "error"
        result.error = str(e)
        return result

    try:
        wb = load_workbook(input_file, data_only=True)
    except Exception as e:
        result.status = "error"
        result.error = f"Failed to load Excel file: {e}"
        return result

    richdata_images = extract_richdata_images(input_file)
    wps_images = extract_wps_cellimages(input_file)

    result.total_sheets = len(wb.sheetnames)

    # Launch Playwright browser once and reuse across all sheets
    pw_ctx = None
    browser = None
    use_playwright = False
    try:
        from playwright.sync_api import sync_playwright
        pw_ctx = sync_playwright().start()
        browser = pw_ctx.chromium.launch()
        use_playwright = True
        print("Playwright browser launched (reusing across sheets)")
    except Exception as e:
        print(f"Playwright not available ({e}), falling back to per-sheet converter")

    try:
        for i, sheet_name in enumerate(wb.sheetnames):
            print(f"Processing {i+1}/{result.total_sheets}: {sheet_name}")

            sheet = wb[sheet_name]
            safe_name = sanitize_filename(sheet_name)

            cell_vm_map = extract_cell_richdata_mapping(input_file, sheet_name)
            dispimg_cells = extract_dispimg_cells(input_file, i)

            # Auto-segment the sheet
            segments = auto_segment(sheet, max_rows)
            n_segments = len(segments)

            for seg_idx, (seg_start, seg_end) in enumerate(segments):
                t0 = time.time()

                # File naming: single segment keeps old naming, multiple segments get _partN
                if n_segments == 1:
                    png_name = f"{safe_name}.png"
                    html_name = f"{safe_name}.html"
                    label = sheet_name
                else:
                    part = seg_idx + 1
                    png_name = f"{safe_name}_part{part}.png"
                    html_name = f"{safe_name}_part{part}.html"
                    label = f"{sheet_name} (part {part}/{n_segments}, rows {seg_start}-{seg_end})"

                sheet_result = SheetScreenshot(
                    sheet_name=label,
                    output_file=png_name,
                )

                try:
                    html_content = sheet_to_html(
                        sheet, sheet_name, wb, input_file,
                        richdata_images, cell_vm_map, wps_images, dispimg_cells,
                        start_row=seg_start, end_row=seg_end,
                        show_title=(seg_idx == 0),
                    )

                    html_file = os.path.join(output_dir, html_name)
                    with open(html_file, 'w', encoding='utf-8') as f:
                        f.write(html_content)

                    png_file = os.path.join(output_dir, png_name)
                    if use_playwright:
                        success = convert_html_to_image_playwright(html_file, png_file, browser=browser)
                    else:
                        success = convert_to_image(html_file, png_file)

                    elapsed = time.time() - t0
                    if success and os.path.exists(png_file):
                        size_kb = os.path.getsize(png_file) / 1024
                        sheet_result.success = True
                        result.success_count += 1
                        print(f"  [{seg_idx+1}/{n_segments}] {label} -> {png_name} ({size_kb:.0f}KB, {elapsed:.1f}s)")
                    else:
                        sheet_result.success = False
                        sheet_result.error = "PNG file was not created"
                        result.failed_count += 1
                        print(f"  FAILED: {label} ({elapsed:.1f}s)")

                except Exception as e:
                    elapsed = time.time() - t0
                    err_name = type(e).__name__
                    sheet_result.success = False
                    sheet_result.error = f"{err_name}: {e}"
                    result.failed_count += 1
                    print(f"  {err_name} ({elapsed:.1f}s): {e}")

                result.sheets.append(sheet_result)

    finally:
        if browser:
            try:
                browser.close()
            except Exception:
                pass
        if pw_ctx:
            try:
                pw_ctx.stop()
            except Exception:
                pass

    if result.failed_count == 0:
        result.status = "success"
    elif result.success_count > 0:
        result.status = "partial"
    else:
        result.status = "error"

    return result


def print_report(result: ScreenshotResult):
    """Print the export report"""
    print("\n" + "=" * 60)
    print("Excel Screenshot Export Report")
    print("=" * 60)
    print(f"Input file: {result.input_file}")
    print(f"Output directory: {result.output_directory}")
    print(f"Total sheets: {result.total_sheets}")
    print(f"Successful: {result.success_count}")
    print(f"Failed: {result.failed_count}")
    print("-" * 60)

    if result.error:
        print(f"\nERROR: {result.error}")

    if result.sheets:
        print("\nSheet Export Details:")
        for sheet in result.sheets:
            status = "OK" if sheet.success else "FAIL"
            print(f"  [{status}] {sheet.sheet_name}")
            if sheet.success:
                print(f"       -> {sheet.output_file}")
            elif sheet.error:
                print(f"       Error: {sheet.error}")

    print("=" * 60)

    if result.status == "success":
        print(f"SUCCESS: All screenshots exported to {result.output_directory}")
    elif result.status == "partial":
        print(f"PARTIAL: {result.success_count} of {result.success_count + result.failed_count} exported")
    else:
        print("FAILED: Export failed")

    print("=" * 60)


def main():
    parser = argparse.ArgumentParser(
        description="Export Excel sheets to PNG images using HTML rendering",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""
Examples:
  python excel_screenshot.py report.xlsx
  python excel_screenshot.py report.xlsx ./output
  python excel_screenshot.py report.xlsx ./output --max-rows 40

Requirements (one of the following):
  - Playwright: pip install playwright && playwright install chromium
  - wkhtmltopdf: brew install wkhtmltopdf (macOS) or apt install wkhtmltopdf (Linux)
Optional:
  - Pillow: pip install Pillow  (for auto-crop whitespace)
        """
    )
    parser.add_argument("input_file", help="Path to the Excel file (.xlsx)")
    parser.add_argument("output_dir", help="Output directory")
    parser.add_argument("--max-rows", type=int, default=MAX_ROWS_PER_SEGMENT,
                        help=f"Max rows per screenshot segment (default: {MAX_ROWS_PER_SEGMENT})")

    args = parser.parse_args()

    if not os.path.exists(args.input_file):
        print(f"Error: File not found: {args.input_file}")
        sys.exit(1)

    input_name = Path(args.input_file).stem
    output_dir = get_output_directory(input_name, args.output_dir)

    print(f"Excel Screenshot Tool")
    print(f"Input: {args.input_file}")
    print(f"Output: {output_dir}")
    print(f"Max rows per segment: {args.max_rows}")
    print("-" * 40)

    result = take_screenshots(args.input_file, output_dir, args.max_rows)
    print_report(result)

    sys.exit(0 if result.status == "success" else 1)


if __name__ == "__main__":
    main()
